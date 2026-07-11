"""Workbook builder for the selection export.

Pure function of ``ExportData`` -> xlsx bytes; no DB, no HTTP. Sheet set:

  Summary            category x formation rollup (reconciles with the app)
  Assumptions        selection metadata + price deck + caveat
  {F} — meta         per-well table for formation F (PUD/RES only)
  {F} — forecast     ip_day grid: formation AVERAGE block only; each stream
                     as daily rate + 30-day volume, PER 1,000 ft of lateral
                     (anduin type-curve basis)
  Arps params        segmented decline params (source of the 50-yr tail)

Built in regular (in-memory) openpyxl mode — a typical selection is a few
hundred thousand cells. If selections ever grow past ~5M cells (hundreds of
wells x many formations), switch to ``Workbook(write_only=True)`` with
``WriteOnlyCell`` for the bold headers and set freeze_panes/column widths
before appending rows.
"""

from __future__ import annotations

import io
import re

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.api.production import STEP
from app.exports.data import ExportData, WellStream

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Excel caps sheet names at 31 chars and forbids \ / ? * [ ] :.
# " — forecast" is the longest suffix at 11 chars, so 20 lands exactly
# on the cap.
_SHEET_SLUG_MAX = 20
_SHEET_FORBIDDEN_RE = re.compile(r"[\\/?*\[\]:]+")

_BOLD = Font(bold=True)

# Per-well columns on the meta tabs, in order. One place to extend.
# `forecast_end_day` marks where the Novi ML forecast hands off to the
# Arps tail on the forecast tab (the old CSV's `source` column).
META_COLS = (
    "unique_id", "category", "operator", "pad_name", "county", "fp_year",
    "tvd", "md", "ll_ft", "prop_load", "forecast_end_day",
    "oil_eur", "gas_eur", "ngl_eur", "water_eur",
    "oil_ip", "gas_ip", "ngl_ip", "water_ip",
    "npv5", "npv10", "npv15", "npv20", "npv25", "pv10",
    "dc_cost", "dcet_cost", "irr_pct",
    # offset-PDP support family (curated.intel_pdp_support, sql/30) — full family
    # on the export (the six UI filters + the export-only distances/medians/n).
    "pdp_count_1mi", "pdp_count_3mi", "pdp_count_5mi",
    "dist_nearest_ft", "dist_3rd_nearest_ft", "support_lateral_ft_5mi",
    "n_offsets_5mi", "offset_median_eur_ft", "offset_median_cum12m_oil_per_ft",
    "inflation_ratio",
)
_META_INT_COLS = {
    "fp_year", "tvd", "md", "ll_ft", "prop_load", "forecast_end_day",
    "oil_eur", "gas_eur", "ngl_eur", "water_eur",
    "npv5", "npv10", "npv15", "npv20", "npv25", "pv10",
    "dc_cost", "dcet_cost",
    # support counts / distances / footage
    "pdp_count_1mi", "pdp_count_3mi", "pdp_count_5mi",
    "dist_nearest_ft", "dist_3rd_nearest_ft", "support_lateral_ft_5mi",
    "n_offsets_5mi",
}
_META_DEC_COLS = {
    "oil_ip", "gas_ip", "ngl_ip", "water_ip", "irr_pct",
    "offset_median_eur_ft", "offset_median_cum12m_oil_per_ft", "inflation_ratio",
}

SUMMARY_SUMS = (
    "npv5", "npv10", "npv15", "npv20", "npv25",
    "pv5", "pv10", "pv15", "pv20", "pv25", "oil_eur", "gas_eur",
)

ARPS_COLS = (
    "novi_wellname", "formation", "production_stream", "segment",
    "segment_curve_type", "b", "d_nom", "d_eff_secant", "d_eff_tangent",
    "q_start", "q_stop", "terminal_day", "day_start", "day_stop",
)

_STREAMS = ("oil", "gas", "water")


def _sheet_slug(name: str, used: set[str]) -> str:
    """Sanitize a formation name into a unique sheet-name prefix.

    Strips Excel-forbidden chars, truncates to ``_SHEET_SLUG_MAX``, and
    appends ``_2``, ``_3`` … on collision.
    """
    base = _SHEET_FORBIDDEN_RE.sub("_", name).strip()
    base = base[:_SHEET_SLUG_MAX] or "formation"
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = f"{base[: _SHEET_SLUG_MAX - len(suffix)]}{suffix}"
        n += 1
    used.add(candidate)
    return candidate


def _bold_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).font = _BOLD


def _write_summary(ws, data: ExportData) -> None:
    header = ["category", "formation_blueox", "count", *SUMMARY_SUMS]
    ws.append(header)
    _bold_row(ws, 1, len(header))

    groups: dict[tuple, dict] = {}
    for r in data.locations:
        key = (r["category"], r.get("formation_blueox") or "(unmapped)")
        g = groups.setdefault(key, {"count": 0, **{c: 0.0 for c in SUMMARY_SUMS}})
        g["count"] += 1
        for c in SUMMARY_SUMS:
            g[c] += float(r.get(c) or 0.0)
    total = {"count": 0, **{c: 0.0 for c in SUMMARY_SUMS}}
    for (cat, formation), g in sorted(groups.items()):
        ws.append([cat, formation, g["count"], *(round(g[c], 2) for c in SUMMARY_SUMS)])
        total["count"] += g["count"]
        for c in SUMMARY_SUMS:
            total[c] += g[c]
    ws.append(["TOTAL", "", total["count"], *(round(total[c], 2) for c in SUMMARY_SUMS)])
    _bold_row(ws, ws.max_row, len(header))

    for col_idx in range(4, len(header) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].number_format = "#,##0"
        ws.column_dimensions[letter].width = 14
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 8


def _write_assumptions(ws, data: ExportData) -> None:
    pud = sum(1 for r in data.locations if r.get("category") == "PUD")
    res = sum(1 for r in data.locations if r.get("category") == "RES")
    ws.append(["field", "value"])
    _bold_row(ws, 1, 2)
    ws.append(["generated", data.generated_at.strftime("%Y-%m-%d %H:%M UTC")])
    ws.append(["basin", data.basin])
    ws.append(["selection_rule", data.rule])
    ws.append(["included_pud_wells", pud])
    ws.append(["included_res_wells", res])
    ws.append(["pdp_in_selection_excluded_from_workbook", data.pdp_count])
    ws.append(["excluded_formations", ", ".join(data.excluded_formations) or "(none)"])
    ws.append(["manually_culled_wells", data.culled_count])
    ws.append(["forecast_horizon_days", 18250])
    ws.append(["forecast_step_days", STEP])
    ws.append([
        "forecast_basis",
        "PER 1,000 ft of lateral — each well's stream is normalized by its ll_ft, "
        "then averaged across wells (matches the anduin type-curve basis). Wells "
        "with no ll_ft are excluded from the average.",
    ])
    ws.append([
        "forecast_rate_units",
        f"oil BOPD/1000ft · gas MCFD/1000ft · water BWPD/1000ft (avg daily rate per {STEP}-day step)",
    ])
    ws.append([
        "volume_convention",
        f"vol = rate × {STEP} (per {STEP}-day period), also per 1,000 ft of lateral",
    ])
    ws.append([
        "tail_convention",
        "rates beyond a well's forecast_end_day (meta tabs) are evaluated "
        "from its Arps segments (Arps params tab)",
    ])
    ws.append([])

    ws.append(["price_deck (flat)"])
    ws.cell(row=ws.max_row, column=1).font = _BOLD
    for k, v in data.price_deck.items():
        ws.append([k, v])
    ws.append([])

    ws.append([
        "CAVEAT: economics here are Novi's pre-computed values on a single "
        "flat price deck — a screening number, not a valuation. Run your own "
        "model off these inputs."
    ])
    ws.cell(row=ws.max_row, column=1).font = _BOLD

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 36


def _write_meta(
    ws, formation: str, loc_rows: list[dict], streams: list[WellStream], data: ExportData
) -> None:
    end_by_name = {s.name: s.forecast_end_day for s in streams}
    pud = sum(1 for r in loc_rows if r.get("category") == "PUD")
    res = sum(1 for r in loc_rows if r.get("category") == "RES")

    ws.append(["formation_blueox", formation])
    ws.append(["wells", f"{len(loc_rows)} (PUD {pud} / RES {res})"])
    ws.append(["generated", data.generated_at.strftime("%Y-%m-%d %H:%M UTC")])
    for row in (1, 2, 3):
        ws.cell(row=row, column=1).font = _BOLD
    ws.append([])

    ws.append(list(META_COLS))
    _bold_row(ws, 5, len(META_COLS))
    for r in sorted(loc_rows, key=lambda x: x.get("unique_id") or ""):
        row_vals = []
        for col in META_COLS:
            if col == "forecast_end_day":
                row_vals.append(end_by_name.get(r.get("unique_id")))
            else:
                row_vals.append(r.get(col))
        ws.append(row_vals)
        data_row = ws.max_row
        for col_idx, col in enumerate(META_COLS, start=1):
            if col in _META_INT_COLS:
                ws.cell(row=data_row, column=col_idx).number_format = "#,##0"
            elif col in _META_DEC_COLS:
                ws.cell(row=data_row, column=col_idx).number_format = "0.0"

    ws.freeze_panes = "A6"
    ws.column_dimensions["A"].width = 26
    for col_idx in range(2, len(META_COLS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13


def _write_forecast(ws, grid: list[int], streams: list[WellStream]) -> None:
    # headers stay plain (per-1,000-ft basis is the SOP + documented on Assumptions;
    # a "/1000ft" suffix reads as "divide by 1000" and invites double-scaling)
    header = ["ip_day"]
    for stream in _STREAMS:
        header.append(f"AVG {stream}_rate")
        header.append(f"AVG {stream}_vol")
    ws.append(header)
    _bold_row(ws, 1, len(header))

    # Per-1,000-ft type curve (matches the anduin TC basis): normalize each well
    # by its own lateral length, then average across wells. Wells without a usable
    # ll_ft are excluded from the normalized average.
    norm = [(s, 1000.0 / s.ll_ft) for s in streams if s.ll_ft and s.ll_ft > 0]
    denom = len(norm) or 1
    for i, day in enumerate(grid):
        row: list[float | int] = [day]
        for stream in _STREAMS:
            per1000 = sum(getattr(s, stream)[i] * factor for s, factor in norm) / denom
            row.append(per1000)
            row.append(per1000 * STEP)
        ws.append(row)

    ws.freeze_panes = "B2"
    for col_idx in range(2, len(header) + 1):
        # Header alternates rate/vol pairs: rate at even col index, vol odd.
        is_rate = (col_idx % 2) == 0
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].number_format = "0.0000" if is_rate else "0.0"
        ws.column_dimensions[letter].width = 16
    ws.column_dimensions["A"].number_format = "0"
    ws.column_dimensions["A"].width = 10


def _write_arps(ws, data: ExportData) -> None:
    ws.append(list(ARPS_COLS))
    _bold_row(ws, 1, len(ARPS_COLS))
    for r in data.arps_rows:
        ws.append([r.get(c) for c in ARPS_COLS])
    ws.column_dimensions["A"].width = 26
    for col_idx in range(2, len(ARPS_COLS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def build_workbook(data: ExportData) -> bytes:
    wb = Workbook()
    # Workbook ships with a default empty sheet; remove it so Summary is
    # the active sheet on open.
    wb.remove(wb.active)

    _write_summary(wb.create_sheet("Summary"), data)
    _write_assumptions(wb.create_sheet("Assumptions"), data)

    # Formation list from the locations (a PUD/RES well without a forecast
    # still shows on its meta tab; it just has no forecast columns).
    locs_by_formation: dict[str, list[dict]] = {}
    for r in data.locations:
        key = r.get("formation_blueox") or "(unmapped)"
        locs_by_formation.setdefault(key, []).append(r)

    used_slugs: set[str] = set()
    for formation in sorted(locs_by_formation):
        slug = _sheet_slug(formation, used_slugs)
        streams = data.streams_by_formation.get(formation, [])
        _write_meta(
            wb.create_sheet(f"{slug} — meta"),
            formation, locs_by_formation[formation], streams, data,
        )
        _write_forecast(wb.create_sheet(f"{slug} — forecast"), data.grid, streams)

    _write_arps(wb.create_sheet("Arps params"), data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
