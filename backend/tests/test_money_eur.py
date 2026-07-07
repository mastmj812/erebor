"""End-to-end "money test" for the deal export.

A fixed synthetic cohort's EXPORTED per-1,000-ft EUR (oil) must match a pinned
value within 0.5%. This is the tripwire for the next convention-level bug: in
one shot it exercises assemble_export_data, the per-lateral-ft normalization,
the Arps-tail integration, and the workbook writer — reading the number
straight off the forecast sheet that ships to finance.

FINDING (reported, not silently worked around): erebor's forecast tab is the
cohort MEAN per 1,000 ft, not a P50 percentile. Unlike anduin (which emits
P10/P50/P90), erebor has no percentile export today. So this pins the exported
MEAN EUR/1000ft. A true P50 export would be a feature to add — flagged here,
not invented.

Pure builder — no database, no Supabase.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

import pytest
from openpyxl import load_workbook

from app.exports.data import assemble_export_data
from app.exports.xlsx_builder import build_workbook

# Pinned exported MEAN oil EUR per 1,000 ft (bbl / 1000 ft) for the fixed
# cohort below. Re-pin ONLY on a deliberate export-math change.
PINNED_MEAN_OIL_EUR_PER_1000FT: float = 5_436.66
TOLERANCE: float = 0.005  # 0.5%


def _loc(uid: str, ll_ft: int) -> dict[str, Any]:
    return {
        "unique_id": uid, "category": "PUD", "formation_blueox": "WCA_1",
        "operator": "OPCO", "pad_name": "PAD 1", "county": "LOVING",
        "fp_year": 2027, "tvd": 9000, "md": 19000, "ll_ft": ll_ft,
        "prop_load": 2400, "oil_eur": 500000, "gas_eur": 2000000,
        "ngl_eur": 100000, "water_eur": 800000, "oil_ip": 800.0,
        "gas_ip": 2500.0, "ngl_ip": 120.0, "water_ip": 900.0,
        "npv5": 1.4e6, "npv10": 1e6, "npv15": 0.8e6, "npv20": 0.6e6,
        "npv25": 0.5e6, "pv5": 1.6e6, "pv10": 1.2e6, "pv15": 0.9e6,
        "pv20": 0.7e6, "pv25": 0.55e6, "dc_cost": 9e6, "dcet_cost": 1.1e7,
        "irr_pct": 45.0, "wti_price": 70.0, "wti_diff": 2.0, "hh_price": 3.5,
        "ngl_price": 25.0, "hh_diff": 0.3,
    }


def _prod(name: str, ip_day: int, oil: float) -> dict[str, Any]:
    return {"novi_wellname": name, "ip_day": ip_day, "oil": oil,
            "gas": oil * 3.0, "water": oil * 0.5}


def _arps(name: str, stream: str, q_start: float) -> dict[str, Any]:
    return {
        "novi_wellname": name, "production_stream": stream, "segment": 2,
        "segment_curve_type": "exponential", "b": 0.0, "d_nom": 0.4,
        "d_eff_secant": 0.33, "d_eff_tangent": 0.4, "q_start": q_start,
        "q_stop": 1.0, "terminal_day": 18250, "day_start": 120, "day_stop": 18250,
    }


def _fixed_export() -> Any:
    loc_rows = [_loc("WELL A 1H", 10000), _loc("WELL B 1H", 8000)]
    prod_rows = [
        _prod(name, day, q - 10 * i)
        for name, q in (("WELL A 1H", 100.0), ("WELL B 1H", 80.0))
        for i, day in enumerate((30, 60, 90))
    ]
    arps_rows = [
        _arps(name, stream, q)
        for name, q0 in (("WELL A 1H", 50.0), ("WELL B 1H", 40.0))
        for stream, q in (("oil", q0), ("gas", q0 * 3), ("water", q0 * 0.5))
    ]
    return assemble_export_data(
        loc_rows, prod_rows, arps_rows, basin="delaware", rule="intersects",
        excluded_formations=[], culled_count=0,
        generated_at=datetime(2026, 7, 7, tzinfo=timezone.utc),
    )


def _forecast_sheet(wb: Any) -> Any:
    names = [s for s in wb.sheetnames if s.endswith("forecast")]
    assert names, f"no forecast sheet found in {wb.sheetnames}"
    return wb[names[0]]


def test_exported_mean_oil_eur_per_1000ft_is_pinned() -> None:
    wb = load_workbook(BytesIO(build_workbook(_fixed_export())))
    ws = _forecast_sheet(wb)
    header = [c.value for c in ws[1]]
    col = header.index("AVG oil_vol")  # per-1,000-ft monthly-ish volume column
    exported = sum(
        (ws.cell(r, col + 1).value or 0.0) for r in range(2, ws.max_row + 1)
    )
    assert exported == pytest.approx(PINNED_MEAN_OIL_EUR_PER_1000FT, rel=TOLERANCE), (
        f"exported oil EUR/1000ft = {exported:.1f}, pinned "
        f"{PINNED_MEAN_OIL_EUR_PER_1000FT:.1f} (±{TOLERANCE * 100:.1f}%). The "
        f"per-1,000-ft type curve the export ships moved — investigate before "
        f"re-pinning."
    )
