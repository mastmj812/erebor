"""Workbook-shape tests for the xlsx export.

Pure-function tests: fake selection rows feed ``assemble_export_data`` and
``build_workbook`` directly — no DB. The route test monkeypatches the
gather step and exercises filename sanitization end-to-end.
"""

from __future__ import annotations

import re
from datetime import datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.exports.data import assemble_export_data
from app.exports.xlsx_builder import (
    ARPS_COLS,
    META_COLS,
    SUMMARY_SUMS,
    _sheet_slug,
    build_workbook,
)

GENERATED = datetime(2026, 6, 12, 12, 0, tzinfo=timezone.utc)


def _loc(uid: str, cat: str, formation_blueox: str, npv10: float, **kw) -> dict:
    base = {
        "unique_id": uid, "category": cat, "formation_blueox": formation_blueox,
        "operator": "OPCO", "pad_name": "PAD 1", "county": "LOVING",
        "fp_year": 2027, "tvd": 9000, "md": 19000, "ll_ft": 10000,
        "prop_load": 2400, "oil_eur": 500000, "gas_eur": 2000000,
        "ngl_eur": 100000, "water_eur": 800000,
        "oil_ip": 800.0, "gas_ip": 2500.0, "ngl_ip": 120.0, "water_ip": 900.0,
        "npv5": npv10 * 1.4, "npv10": npv10, "npv15": npv10 * 0.8,
        "npv20": npv10 * 0.6, "npv25": npv10 * 0.5,
        "pv5": npv10 * 1.6, "pv10": npv10 * 1.2, "pv15": npv10 * 0.9,
        "pv20": npv10 * 0.7, "pv25": npv10 * 0.55,
        "dc_cost": 9e6, "dcet_cost": 1.1e7, "irr_pct": 45.0,
        "wti_price": 70.0, "wti_diff": 2.0, "hh_price": 3.5,
        "ngl_price": 25.0, "hh_diff": 0.3,
    }
    base.update(kw)
    return base


def _prod(name: str, ip_day: int, oil: float) -> dict:
    return {"novi_wellname": name, "ip_day": ip_day,
            "oil": oil, "gas": oil * 3.0, "water": oil * 0.5}


def _arps(name: str, stream: str, q_start: float) -> dict:
    return {
        "novi_wellname": name, "production_stream": stream, "segment": 2,
        "segment_curve_type": "exponential", "b": 0.0, "d_nom": 0.4,
        "d_eff_secant": 0.33, "d_eff_tangent": 0.4, "q_start": q_start,
        "q_stop": 1.0, "terminal_day": 18250, "day_start": 120, "day_stop": 18250,
    }


@pytest.fixture(scope="module")
def data():
    # formation_blueox codes are already canonical (no case transform). Wells in
    # WCA_1 carry DIFFERENT ll_ft so the per-1,000-ft normalization is testable.
    loc_rows = [
        _loc("WELL A 1H", "PUD", "WCA_1", npv10=1e6, ll_ft=10000),
        _loc("WELL A 2H", "RES", "WCA_1", npv10=2e6, ll_ft=5000),
        _loc("WELL B 1H", "PUD", "BS2_C", npv10=3e6, ll_ft=8000),
        _loc("4230112345", "PDP", "WCA_1", npv10=9e9),
    ]
    prod_rows = []
    for name, q in (("WELL A 1H", 100.0), ("WELL A 2H", 140.0), ("WELL B 1H", 60.0)):
        for i, day in enumerate((30, 60, 90)):
            prod_rows.append(_prod(name, day, q - 10 * i))
    arps_rows = []
    for name in ("WELL A 1H", "WELL A 2H", "WELL B 1H"):
        for stream, q in (("oil", 50.0), ("gas", 150.0), ("water", 25.0)):
            arps_rows.append(_arps(name, stream, q))
    return assemble_export_data(
        loc_rows, prod_rows, arps_rows,
        basin="delaware", rule="intersects",
        excluded_formations=["AVALON"], culled_count=2,
        generated_at=GENERATED,
    )


@pytest.fixture(scope="module")
def wb(data):
    return load_workbook(BytesIO(build_workbook(data)))


def test_assemble_drops_pdp_and_groups_by_blueox(data):
    assert data.pdp_count == 1
    assert all(r["category"] in ("PUD", "RES") for r in data.locations)
    assert set(data.streams_by_formation) == {"BS2_C", "WCA_1"}
    assert [s.name for s in data.streams_by_formation["WCA_1"]] == [
        "WELL A 1H", "WELL A 2H",
    ]
    # ll_ft threads onto the stream for the per-1,000-ft basis.
    assert [s.ll_ft for s in data.streams_by_formation["WCA_1"]] == [10000.0, 5000.0]
    # Grid: forecast days 30/60/90 + tail 120..18240 on 30-day cadence.
    assert data.grid[0] == 30
    assert data.grid[-1] == 18240
    assert all(b - a == 30 for a, b in zip(data.grid, data.grid[1:]))
    # PDP has no arps rows in the fixture, but the filter must hold anyway.
    assert all(r["novi_wellname"] != "4230112345" for r in data.arps_rows)
    assert all(r["formation"] in ("BS2_C", "WCA_1") for r in data.arps_rows)


def test_sheet_names_and_order(wb):
    assert wb.sheetnames == [
        "Summary",
        "Assumptions",
        "BS2_C — meta",
        "BS2_C — forecast",
        "WCA_1 — meta",
        "WCA_1 — forecast",
        "Arps params",
    ]


def test_forecast_sheet_shape_and_math(wb, data):
    ws = wb["WCA_1 — forecast"]
    assert ws.max_row == len(data.grid) + 1
    assert ws.max_column == 1 + 6  # ip_day + AVG {oil,gas,water} x {rate,vol}
    assert ws.freeze_panes == "B2"
    header = [c.value for c in ws[1]]
    assert header == [
        "ip_day",
        "AVG oil_rate", "AVG oil_vol",
        "AVG gas_rate", "AVG gas_vol",
        "AVG water_rate", "AVG water_vol",
    ]

    # Per-1,000-ft basis: each well normalized by its own ll_ft, then averaged.
    # ip_day 30 oil — A1H 100/(10000/1000)=10 ; A2H 140/(5000/1000)=28 ; mean=19.
    row = [c.value for c in ws[2]]
    assert row[0] == 30
    assert row[1] == pytest.approx((100.0 / 10 + 140.0 / 5) / 2)  # == 19.0
    for rate_idx in (1, 3, 5):
        assert row[rate_idx + 1] == pytest.approx(row[rate_idx] * 30)  # vol == rate*30

    # Tail rows: arps exponential starts at q_start (t=0) and decays.
    # oil q_start 50 for both -> A1H 50/10=5 ; A2H 50/5=10 ; mean=7.5.
    tail_row = [c.value for c in ws[5]]  # ip_day 120 == segment day_start
    assert tail_row[0] == 120
    assert tail_row[1] == pytest.approx((50.0 / 10 + 50.0 / 5) / 2)  # == 7.5
    later_row = [c.value for c in ws[20]]
    assert 0 < later_row[1] < 7.5


def test_meta_sheet(wb, data):
    ws = wb["WCA_1 — meta"]
    assert ws.freeze_panes == "A6"
    assert ws["B1"].value == "WCA_1"
    assert ws["B2"].value == "2 (PUD 1 / RES 1)"
    header = [c.value for c in ws[5]]
    assert header == list(META_COLS)
    names = [ws.cell(row=r, column=1).value for r in (6, 7)]
    assert names == ["WELL A 1H", "WELL A 2H"]
    # forecast_end_day column carries the Novi->Arps handoff day.
    fed_col = META_COLS.index("forecast_end_day") + 1
    assert ws.cell(row=6, column=fed_col).value == 90


def test_summary_reconciles_with_meta(wb):
    ws = wb["Summary"]
    header = [c.value for c in ws[1]]
    assert header == ["category", "formation_blueox", "count", *SUMMARY_SUMS]
    rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    cats = {r[0] for r in rows}
    assert "PDP" not in cats
    npv10_col = header.index("npv10")
    by_form: dict[str, float] = {}
    total_row = rows[-1]
    assert total_row[0] == "TOTAL"
    for r in rows[:-1]:
        by_form[r[1]] = by_form.get(r[1], 0.0) + r[npv10_col]
    # Reconcile against the meta tabs' npv10 column.
    meta_npv10_col = META_COLS.index("npv10") + 1
    for formation, expected in by_form.items():
        meta = wb[f"{formation} — meta"]
        got = sum(
            meta.cell(row=r, column=meta_npv10_col).value or 0.0
            for r in range(6, meta.max_row + 1)
        )
        assert got == pytest.approx(expected)
    assert total_row[npv10_col] == pytest.approx(sum(by_form.values()))
    assert total_row[2] == 3


def test_assumptions_sheet(wb):
    ws = wb["Assumptions"]
    fields = {row[0].value: row[1].value for row in ws.iter_rows() if row[0].value}
    assert fields["basin"] == "delaware"
    assert fields["pdp_in_selection_excluded_from_workbook"] == 1
    assert fields["excluded_formations"] == "AVALON"
    assert fields["manually_culled_wells"] == 2
    assert fields["wti_price"] == 70.0
    # the per-1,000-ft forecast convention is documented for the reader
    assert fields["forecast_basis"].startswith("PER 1,000 ft")
    assert "/1000ft" in fields["forecast_rate_units"]


def test_arps_sheet(wb, data):
    ws = wb["Arps params"]
    assert [c.value for c in ws[1]] == list(ARPS_COLS)
    assert ws.max_row == len(data.arps_rows) + 1


def test_sheet_slug_cap_and_collisions():
    used: set[str] = set()
    long_name = "THIRD BONE SPRING CARBONATE LOWER"
    slug = _sheet_slug(long_name, used)
    assert len(slug) <= 20
    assert len(f"{slug} — forecast") <= 31
    slug2 = _sheet_slug(long_name, used)
    assert slug2 != slug and slug2.endswith("_2")
    assert len(slug2) <= 20
    # Runs of forbidden chars collapse to a single underscore.
    assert _sheet_slug("BAD[]:/\\NAME?", set()) == "BAD_NAME_"


# ---------------------------- route tests ----------------------------


@pytest.fixture()
def client(monkeypatch, data):
    from fastapi.testclient import TestClient

    import app.api.export as export_mod
    from app.db import get_session
    from app.main import app

    monkeypatch.setattr(export_mod, "gather_export_data", lambda session, body: data)
    app.dependency_overrides[get_session] = lambda: None
    try:
        yield TestClient(app)
    finally:
        app.dependency_overrides.pop(get_session, None)


_BODY = {"aoi": {"type": "Polygon", "coordinates": []}, "basin": "delaware"}


def test_route_returns_xlsx(client):
    r = client.post("/api/export", json=_BODY)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(BytesIO(r.content))
    assert wb.sheetnames[0] == "Summary"


def test_route_default_filename(client):
    r = client.post("/api/export", json=_BODY)
    m = re.search(r'filename="([^"]+)"', r.headers["content-disposition"])
    assert m and re.fullmatch(r"erebor_delaware_\d{4}-\d{2}-\d{2}\.xlsx", m.group(1))


def test_route_sanitizes_filename(client):
    r = client.post(
        "/api/export", json={**_BODY, "filename": '../evil:na"me?\r\n.zip'}
    )
    m = re.search(r'filename="([^"]+)"', r.headers["content-disposition"])
    assert m
    name = m.group(1)
    assert name.endswith(".xlsx")
    assert re.fullmatch(r"[A-Za-z0-9 ._()-]+\.xlsx", name)
    assert ".." not in name


def test_route_user_filename(client):
    r = client.post("/api/export", json={**_BODY, "filename": "braveheart DSU 12"})
    m = re.search(r'filename="([^"]+)"', r.headers["content-disposition"])
    assert m and m.group(1) == "braveheart DSU 12.xlsx"
