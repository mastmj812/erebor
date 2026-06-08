"""Phase 0 discovery: peek CSV headers + sample rows (without loading the 2.86 GB forecast),
and list xlsx sheets/headers. Read-only. Goal: reconcile shapefile 'Unique ID' with CSV
'novi_wellname', and see what the Data Download xlsx holds."""
import csv, glob, itertools, os, sys

ROOTS = {
    "delaware": r"C:\Users\MichaelMast\Blue Ox Resources\Engineering - General\Novi Intelligence\Delaware\3Q25",
    "midland":  r"C:\Users\MichaelMast\Blue Ox Resources\Engineering - General\Novi Intelligence\Midland\3Q25",
}
KINDS = {  # keyword in filename -> label
    "analytics": "ANALYTICS",
    "arps":      "ARPS",
    "forecast":  "FORECAST",
}


def find_csv(root, kw):
    hits = []
    for p in glob.glob(os.path.join(root, "**", "*.csv"), recursive=True):
        if kw in os.path.basename(p).lower():
            hits.append(p)
    return hits


def head_csv(path, n=3):
    size_mb = round(os.path.getsize(path) / 1e6, 1)
    print(f"  {os.path.basename(path)}  ({size_mb} MB)")
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as fh:
        rdr = csv.reader(fh)
        rows = list(itertools.islice(rdr, n + 1))
    if not rows:
        print("    (empty)")
        return
    hdr = rows[0]
    print(f"    header ({len(hdr)}): {hdr}")
    for i, r in enumerate(rows[1:], 1):
        print(f"    row{i}: {r}")


def peek_xlsx(root):
    import openpyxl
    for p in glob.glob(os.path.join(root, "**", "*.xlsx"), recursive=True):
        size_mb = round(os.path.getsize(p) / 1e6, 1)
        print(f"\n  XLSX {os.path.basename(p)}  ({size_mb} MB)")
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        except Exception as e:
            print(f"    !! load failed: {e}")
            continue
        print(f"    sheets: {wb.sheetnames}")
        for ws in wb.worksheets:
            try:
                first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            except Exception as e:
                first = f"<err {e}>"
            print(f"    sheet '{ws.title}' header: {first}")
        wb.close()


if __name__ == "__main__":
    do_xlsx = "--xlsx" in sys.argv
    for basin, root in ROOTS.items():
        print(f"\n{'#'*90}\n# {basin.upper()}\n{'#'*90}")
        for kw, label in KINDS.items():
            print(f"\n[{label}]")
            for p in find_csv(root, kw):
                head_csv(p)
        if do_xlsx:
            peek_xlsx(root)
